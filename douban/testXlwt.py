# -*- coding = utf-8 -*-
# coding=gbk

#将数据保存为EXCEL格式
import xlwt
"""
workbook = xlwt.Workbook(encoding="utf-8") #创建workbook对象
worksheet = workbook.add_sheet('sheet1')    #创建工作表
worksheet.write(0,0,'')                   #写入内容，第一个参数表示行，第二个表示列，第三个为参数
workbook.save('student.xls')                   #保存数据
"""

#将99乘法表写入Excel格式
from openpyxl import Workbook
wb = Workbook()
ws = wb.active      #创建默认工作表
ws.title = "99乘法表"

for i in range(1,10):
    for j in range(1,i+1):
        cell = ws.cell(row = i, column = j)
        cell_value = str(i) + "×" + str(j) + "=" + str(i*j)
        cell.value =  cell_value
        #ws.write(i,j, "%d*%d=%d"%(i+1,j+1,(i+1)*(j+1)))
wb.save("九九乘法表.xlsx")

print("恭喜你，表格创建成功了！！！")